import openpyxl
import spacy
from negspacy.negation import Negex
import logging
import ataxaid

logging.getLogger('root').setLevel(logging.ERROR)

nlp_model = None

def wrap_in_span(text:str) -> spacy.tokens.Span:
    global nlp_model
    if nlp_model is None:
        logging.debug('Loading en_ner_bc5cdr_md...')
        #nlp_model = spacy.load("en_ner_bc5cdr_md")
        #nlp_model.add_pipe("negex", config={"ent_types":["DISEASE"]})
        #nlp_model = spacy.load("en_ner_bionlp13cg_md")
        nlp_model = spacy.load("en_core_sci_scibert")
        nlp_model.add_pipe("negex", config={"ent_types":["ENTITY"]})
    
    doc = nlp_model(text)
    # Definir el inicio y el final del Span dentro del Doc
    # Por ejemplo, para envolver "a span" en un Span
    inicio = 0
    final = len(doc)

    # Convertir las posiciones de caracteres a tokens
    inicio_token = len([token for token in doc.text[:inicio]])
    final_token = inicio_token + len([token for token in doc.text[inicio:final]])

    # Crear el Span
    # Los índices de inicio y final deben ser en términos de tokens, no de caracteres
    span = spacy.tokens.Span(doc, inicio_token, final_token, label="ENTITY")
    return span

# Ruta del archivo Excel
archivo_excel = '/Users/paula/Library/CloudStorage/OneDrive-UniversidadedaCoruña/TESIS/PAPERS/paper1/Libro_original.xlsx'

# Cargar el workbook Excel
wb = openpyxl.load_workbook(archivo_excel)

informes_mix = {}

# Iterar a través de todas las hojas en el workbook
for nombre_hoja in wb.sheetnames:
    print(f"Leyendo hoja: {nombre_hoja}")
    hoja = wb[nombre_hoja]

    resultados_informe = []

    # Leer las celdas de la columna 'AG', omitiendo las dos primeras filas
    # Asumimos que 'AG' es la columna 33
    for fila in hoja.iter_rows(min_row=3, min_col=33, max_col=33, values_only=True):
        if fila[0] is not None and fila[0].strip()!='':
            resultados_informe.append(wrap_in_span(fila[0]))
    informes_mix[nombre_hoja] = resultados_informe

# No olvides cerrar el workbook si ya no lo necesitas
wb.close()

for informe in informes_mix:
    terminos = informes_mix[informe]
    print(informe)

    for termino in terminos:
        matches = ataxaid.get_HPO_matches(termino)

        if len(matches) > 0:
            m = matches[0]
            print(f'{termino.text};{m.HPO.name};{m.HPO.id}')
        else:
            print(f'{termino.text}')
